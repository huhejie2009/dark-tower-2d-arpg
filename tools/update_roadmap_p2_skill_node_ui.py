from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


PROJECT_ROOT = Path(r"H:\GODOT_PROJECT\dark-tower-2d-arpg")
SOURCE = PROJECT_ROOT / "docs" / "planning" / "2026-06-08-dark-tower-2d-arpg-production-roadmap-updated-p2-skill-node-growth-cn-fixed.xlsx"
TARGET = PROJECT_ROOT / "docs" / "planning" / "2026-06-08-dark-tower-2d-arpg-production-roadmap-updated-p2-skill-node-ui.xlsx"

GODOT = r"C:\Users\huhej\OneDrive\桌面\Godot_v4.6.2-stable_win64_console.exe"
PROJECT = r"H:\GODOT_PROJECT\dark-tower-2d-arpg"

OVERVIEW_ACCEPTANCE = (
	"攻击节奏、受击反馈、敌人行为 profile、掉落质量曲线、装备推荐、掉落通知、"
	"背包推荐标签、装备对比摘要、10 分钟刷宝目标检查、三节点技能成长接口、"
	"背包技能节点列表 UI 已接入；玩家能读懂刷宝、换装和成长收益，并可按 QA 指标验收"
)

PHASE_ACCEPTANCE = (
	"攻击手感、受击反馈、敌人行为、掉落质量、装备推荐、背包可读性、技能升级预览、"
	"装备对比摘要、10 分钟刷宝验收机制、最小三节点技能成长接口和背包技能节点列表 UI "
	"第一轮已落地；下一步需要真实人工试玩数据和 P3 楼层内容节奏"
)

TASK_NAME = "刷宝目标、装备推荐、对比摘要与技能成长 UI 可读性"
TASK_STATUS = "第一轮完成+技能节点列表 UI"
TASK_ACCEPTANCE = (
	"InventoryEquipmentWindow 接入 SkillNodeGrowthService / SkillUpgradePreviewService 的三节点预览；"
	"背包 Skills 区显示 basic_attack_training / vitality_training / precision_training 紧凑列表；"
	"支持选择节点、查看收益、升级选中节点；旧基础攻击升级按钮保持兼容；不引入代码生成素材"
)
TASK_VALIDATION = (
	"regression_inventory_skill_node_list_ui.gd；regression_skill_node_growth_service.gd；"
	"regression_skill_upgrade_readability.gd；regression_inventory_skill_upgrade_ui.gd；"
	"FOCUSED_P2_SKILL_NODE_UI_REGRESSION_OK；完整回归"
)


def focus_command() -> str:
	tests = [
		"regression_inventory_skill_node_list_ui.gd",
		"regression_inventory_skill_upgrade_ui.gd",
		"regression_skill_node_growth_service.gd",
		"regression_skill_upgrade_readability.gd",
		"regression_skill_point_basic_attack_upgrade.gd",
		"regression_player_experience_leveling.gd",
		"regression_hud_level_experience_contract.gd",
		"regression_p2_loot_loop_acceptance_service.gd",
	]
	test_items = ",".join(f"'res://tests/regression/{name}'" for name in tests)
	return f"$godot = '{GODOT}'; $project = '{PROJECT}'; $tests = @({test_items})"


def replace_p2_acceptance(ws, value: str) -> None:
	for row in range(1, ws.max_row + 1):
		if ws.cell(row, 1).value == "P2":
			ws.cell(row, 6).value = value
			return
	raise RuntimeError(f"P2 row not found in {ws.title}")


def replace_t009(ws) -> None:
	for row in range(1, ws.max_row + 1):
		if ws.cell(row, 1).value == "T-009":
			ws.cell(row, 5).value = TASK_NAME
			ws.cell(row, 6).value = TASK_STATUS
			ws.cell(row, 7).value = TASK_ACCEPTANCE
			ws.cell(row, 8).value = TASK_VALIDATION
			for col in range(1, 9):
				ws.cell(row, col).fill = PatternFill("solid", fgColor="E2F0D9")
			return
	raise RuntimeError("T-009 row not found")


def append_or_replace_command(ws) -> None:
	title = "P2 技能节点列表 UI 聚焦回归"
	for row in range(1, ws.max_row + 1):
		if ws.cell(row, 1).value == title:
			ws.cell(row, 2).value = focus_command()
			return
	row = ws.max_row + 1
	ws.cell(row, 1).value = title
	ws.cell(row, 2).value = focus_command()
	for col in range(1, 3):
		above = ws.cell(row - 1, col)
		cell = ws.cell(row, col)
		if above.has_style:
			cell.font = copy(above.font)
			cell.fill = copy(above.fill)
			cell.border = copy(above.border)
			cell.alignment = copy(above.alignment)
			cell.number_format = above.number_format


def collect_bad_cells(wb) -> list[str]:
	bad: list[str] = []
	for ws in wb.worksheets:
		for row in ws.iter_rows():
			for cell in row:
				if isinstance(cell.value, str) and "??" in cell.value:
					bad.append(f"{ws.title}!{cell.coordinate}")
	return bad


def main() -> None:
	wb = load_workbook(SOURCE)
	replace_p2_acceptance(wb.worksheets[0], OVERVIEW_ACCEPTANCE)
	wb.worksheets[0]["H7"] = 0.83
	replace_p2_acceptance(wb.worksheets[1], PHASE_ACCEPTANCE)
	replace_t009(wb.worksheets[2])
	append_or_replace_command(wb.worksheets[7])
	bad = collect_bad_cells(wb)
	if bad:
		raise RuntimeError("ROADMAP contains mojibake cells: " + ", ".join(bad[:20]))
	wb.save(TARGET)
	print(TARGET)


if __name__ == "__main__":
	main()
