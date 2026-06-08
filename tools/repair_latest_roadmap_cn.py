from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


PROJECT_ROOT = Path(r"H:\GODOT_PROJECT\dark-tower-2d-arpg")
SOURCE = PROJECT_ROOT / "docs" / "planning" / "2026-06-08-dark-tower-2d-arpg-production-roadmap-updated-p2-skill-node-growth.xlsx"
TARGET = PROJECT_ROOT / "docs" / "planning" / "2026-06-08-dark-tower-2d-arpg-production-roadmap-updated-p2-skill-node-growth-cn-fixed.xlsx"


OVERVIEW_ACCEPTANCE = (
	"攻击节奏、受击反馈、敌人行为 profile、掉落质量曲线、装备推荐、掉落通知、"
	"背包推荐标签、装备对比摘要、10 分钟刷宝目标检查、三节点技能成长接口已接入；"
	"玩家能读懂刷宝、换装和成长收益，并可按 QA 指标验收"
)

PHASE_ACCEPTANCE = (
	"攻击手感、受击反馈、敌人行为、掉落质量、装备推荐、背包可读性、技能升级预览、"
	"装备对比摘要、10 分钟刷宝验收机制和最小三节点技能成长接口第一轮已落地；"
	"下一步需要真实人工试玩数据和 P3 楼层内容节奏"
)

TASK_NAME = "刷宝目标、装备推荐、对比摘要与技能成长可读性"
TASK_STATUS = "第一轮完成+三节点技能接口"
TASK_ACCEPTANCE = (
	"SkillNodeGrowthService 提供 basic_attack_training / vitality_training / precision_training "
	"三个成长节点；SkillUpgradePreviewService 统一输出节点预览；旧基础攻击升级 API 保持兼容；"
	"P2LootLoopAcceptanceService 保留 10 分钟刷宝目标检查；不引入代码生成素材"
)
TASK_VALIDATION = (
	"regression_skill_node_growth_service.gd；regression_skill_upgrade_readability.gd；"
	"regression_skill_point_basic_attack_upgrade.gd；regression_p2_loot_loop_acceptance_service.gd；"
	"FOCUSED_P2_SKILL_NODE_GROWTH_REGRESSION_OK；完整回归"
)

GODOT = r"C:\Users\huhej\OneDrive\桌面\Godot_v4.6.2-stable_win64_console.exe"
PROJECT = r"H:\GODOT_PROJECT\dark-tower-2d-arpg"

FOCUSED_COMMANDS = {
	"P2 装备对比摘要聚焦回归": [
		"regression_equipment_compare_summary_service.gd",
		"regression_equipment_compare_text.gd",
		"regression_equipment_score_service.gd",
		"regression_inventory_item_visual_metadata.gd",
		"regression_inventory_equipment_selection_actions.gd",
		"regression_inventory_recommendation_tags.gd",
		"regression_equipment_recommendation_service.gd",
	],
	"P2 三节点技能成长聚焦回归": [
		"regression_skill_node_growth_service.gd",
		"regression_skill_upgrade_readability.gd",
		"regression_skill_point_basic_attack_upgrade.gd",
		"regression_inventory_skill_upgrade_ui.gd",
		"regression_player_experience_leveling.gd",
		"regression_hud_level_experience_contract.gd",
		"regression_p2_loot_loop_acceptance_service.gd",
	],
	"P2 10分钟刷宝验收聚焦回归": [
		"regression_p2_loot_loop_acceptance_service.gd",
		"regression_loot_quality_service.gd",
		"regression_loot_quality_rules.gd",
		"regression_inventory_recommendation_tags.gd",
		"regression_equipment_compare_summary_service.gd",
		"regression_skill_upgrade_readability.gd",
		"regression_ten_floor_stability.gd",
	],
}


def _contains_mojibake(value: object) -> bool:
	if not isinstance(value, str):
		return False
	return "??" in value or "????????" in value


def _repair_p2_row(ws, acceptance_text: str) -> None:
	for row in range(1, ws.max_row + 1):
		if ws.cell(row, 1).value == "P2":
			ws.cell(row, 6).value = acceptance_text
			return
	raise RuntimeError(f"P2 row not found in {ws.title}")


def _repair_t009_row(ws) -> None:
	for row in range(1, ws.max_row + 1):
		if ws.cell(row, 1).value == "T-009":
			ws.cell(row, 5).value = TASK_NAME
			ws.cell(row, 6).value = TASK_STATUS
			ws.cell(row, 7).value = TASK_ACCEPTANCE
			ws.cell(row, 8).value = TASK_VALIDATION
			for col in range(1, 9):
				ws.cell(row, col).fill = PatternFill("solid", fgColor="E2F0D9")
			return
	raise RuntimeError("T-009 row not found")


def _build_focus_command(tests: list[str]) -> str:
	test_items = ",".join(f"'res://tests/regression/{name}'" for name in tests)
	return (
		f"$godot = '{GODOT}'; $project = '{PROJECT}'; "
		f"$tests = @({test_items})"
	)


def _repair_focus_commands(ws) -> None:
	start_row = 13
	for index, (title, tests) in enumerate(FOCUSED_COMMANDS.items()):
		row = start_row + index
		ws.cell(row, 1).value = title
		ws.cell(row, 2).value = _build_focus_command(tests)


def _collect_bad_cells(wb) -> list[str]:
	bad: list[str] = []
	for ws in wb.worksheets:
		for row in ws.iter_rows():
			for cell in row:
				if _contains_mojibake(cell.value):
					bad.append(f"{ws.title}!{cell.coordinate}")
	return bad


def main() -> None:
	wb = load_workbook(SOURCE)
	_repair_p2_row(wb.worksheets[0], OVERVIEW_ACCEPTANCE)
	wb.worksheets[0]["H7"] = 0.80
	_repair_p2_row(wb.worksheets[1], PHASE_ACCEPTANCE)
	_repair_t009_row(wb.worksheets[2])
	_repair_focus_commands(wb.worksheets[7])
	bad_cells = _collect_bad_cells(wb)
	if bad_cells:
		raise RuntimeError("ROADMAP still contains mojibake cells: " + ", ".join(bad_cells[:20]))
	wb.save(TARGET)
	print(TARGET)


if __name__ == "__main__":
	main()
